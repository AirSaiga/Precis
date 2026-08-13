"""
L02 隐藏行为测试 — 分块加载 v1（中间帧释放 + 快照时机）与 v2（Excel openpyxl read_only 重写）。

本文件由 challenges/L02-evol-chunked-streaming/verify.py 临时注入到
backend/tests/unit/test_l02_v1v2.py 后以 pytest 运行，verify 完成后清理。

只验证可观察行为，不依赖实现细节：
- v1: 快照带真实时间戳与进程级 RSS；分块执行期间快照覆盖合并窗口（时机修正）
- v2: Excel 分块重写后行号全局连续、列名/分块/尾部空行语义与既有行为一致
"""

from __future__ import annotations

import time
from unittest.mock import MagicMock

import pytest

from app.shared.domain.constraints import NotNullConstraint
from app.shared.domain.data_types import IntegerType, StringType
from app.shared.domain.dataset_schema import ColumnSchema, DataSetSchema, TableSchema
from app.shared.services.validation.chunked_loader import ChunkedDataLoader
from app.shared.services.validation.executor import (
    ValidationExecutor,
    ValidationOptions,
)
from app.shared.services.validation.memory_monitor import MemoryMonitor
from app.shared.services.validation.resolver import DataSourceResolver

# ============================================================================
# v1-a: memory_monitor 快照修正
# ============================================================================


class TestMemoryMonitorFix:
    def test_snapshot_records_real_timestamp(self, tmp_path):
        """快照必须记录真实的采集时间戳（而不是默认的 0）。"""
        f = tmp_path / "data.csv"
        f.write_text("a\n1\n", encoding="utf-8")
        monitor = MemoryMonitor()
        before = time.time()
        snapshot = monitor.take_snapshot(str(f))
        after = time.time()
        assert snapshot.timestamp > 0
        assert before - 1 <= snapshot.timestamp <= after + 1

    def test_snapshot_records_process_rss(self, tmp_path):
        """快照必须记录进程级内存占用（RSS），并随 to_dict 输出。"""
        pytest.importorskip("psutil")
        f = tmp_path / "data.csv"
        f.write_text("a\n1\n", encoding="utf-8")
        monitor = MemoryMonitor()
        snapshot = monitor.take_snapshot(str(f))
        assert snapshot.process_memory_mb is not None
        assert snapshot.process_memory_mb > 0
        d = snapshot.to_dict()
        assert "process_memory_mb" in d
        assert "timestamp" in d


# ============================================================================
# v1-b: 分块执行期间的快照时机（覆盖合并窗口）+ 结果不变
# ============================================================================


class TestExecutorSnapshotTiming:
    def test_snapshots_cover_merge_window(self, tmp_path):
        """真实分块链路执行后：快照应覆盖加载期与合并窗口前后（≥3 次），
        时间戳真实且单调不减；同时分块处理结果必须保持不变（无错误、行数一致）。"""
        csv_path = tmp_path / "big.csv"
        lines = ["id,name"] + [f"{i},n{i}" for i in range(350)]
        csv_path.write_text("\n".join(lines), encoding="utf-8")

        schema = DataSetSchema(
            tables={
                "main": TableSchema(
                    id="main",
                    name="main",
                    columns=[
                        ColumnSchema(name="id", id="id", data_type=IntegerType()),
                        ColumnSchema(name="name", id="name", data_type=StringType()),
                    ],
                ),
            },
            constraints=[NotNullConstraint(table="main", column="id")],
        )

        executor = ValidationExecutor.__new__(ValidationExecutor)
        executor.project_root = "D:\\project"
        executor.loaded_project = MagicMock()
        executor.loaded_project.loading_errors = []
        executor.loaded_project.warnings = []
        executor.dataset_schema = schema
        executor.settings = MagicMock()
        executor.manifest = MagicMock()
        executor.allow_unsafe_eval = None
        executor._schema_by_id = {}
        executor._resolve_allow_unsafe_eval = lambda options: False

        schema_file = MagicMock()
        schema_file.source_file = str(csv_path)
        schema_file.sheet_name = None
        schema_file.header_row = 0
        schema_file.source_config = {"delimiter": ","}

        resolver = DataSourceResolver("D:\\project", MagicMock(), {})
        resolver.resolve_first_data_source = lambda: str(tmp_path)
        resolver.resolve_source_path = lambda data_directory, sf: (str(csv_path), None)
        executor._resolver = resolver

        monitor = MemoryMonitor(chunk_threshold_mb=0.001, chunk_rows=100)
        executor._memory_monitor = monitor
        loader = ChunkedDataLoader(
            resolver, schema, {"main": schema_file}, MagicMock(), memory_monitor=monitor
        )
        executor._get_chunked_loader = lambda options: loader

        result = {
            "raw_datasets": {},
            "parsed_datasets": {},
            "errors": [],
            "loading_errors": [],
            "duration_ms": 0,
            "timeout_occurred": False,
            "validation_details": {"format_checks": [], "constraint_checks": []},
            "chunked_mode": False,
            "memory_info": {},
            "warnings": [],
        }
        result = executor._execute_chunked(
            str(tmp_path),
            ValidationOptions(
                timeout_seconds=300, chunk_threshold_mb=0.001, chunk_rows=100
            ),
            time.monotonic(),
            result,
        )

        # 结果不变：中间帧释放/即时合并不得改变校验结果
        assert result["errors"] == [], f"不应有错误: {result['errors']}"
        assert result["parsed_datasets"]["main"].shape[0] == 350

        # 快照时机：加载期（loader 1 次）+ 合并窗口前后（≥2 次）
        assert len(monitor.snapshots) >= 3, f"快照数不足: {len(monitor.snapshots)}"
        timestamps = [s.timestamp for s in monitor.snapshots]
        assert all(t > 0 for t in timestamps)
        assert timestamps == sorted(timestamps)


# ============================================================================
# v2: Excel 分块重写的行为等价（行号全局连续 / 列名 / 分块 / 尾部空行）
# ============================================================================


class TestExcelChunkedRewrite:
    @staticmethod
    def _write_excel(path, rows):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r, row in enumerate(rows, start=1):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=v)
        wb.save(path)
        wb.close()

    def test_multi_chunk_continuous_row_index(self, tmp_path):
        """250 行 × chunk 100 → 3 块，行号全局连续，列名与数据正确。"""
        header = ["id", "name"]
        data = [[i, f"n{i}"] for i in range(250)]
        self._write_excel(tmp_path / "t.xlsx", [header] + data)

        loader = ChunkedDataLoader.__new__(ChunkedDataLoader)
        chunks = loader._load_excel_chunked(
            str(tmp_path / "t.xlsx"), "Sheet1", header_row=0, chunk_size=100
        )

        assert len(chunks) == 3
        assert list(chunks[0].columns) == ["id", "name"]
        assert list(chunks[1].columns) == ["id", "name"]
        assert chunks[0].index.tolist() == list(range(0, 100))
        assert chunks[1].index.tolist() == list(range(100, 200))
        assert chunks[2].index.tolist() == list(range(200, 250))
        assert chunks[1].iloc[0]["id"] == 100
        assert chunks[2].iloc[-1]["name"] == "n249"

    def test_header_row_offset(self, tmp_path):
        """表头不在首行（header_row=2）时，行号仍从 0 起全局连续。"""
        rows = [
            ["报表元数据", None],
            ["生成日期", "2026-08-13"],
            ["id", "amount"],
        ]
        data = [[i, i * 2] for i in range(120)]
        self._write_excel(tmp_path / "t.xlsx", rows + data)

        loader = ChunkedDataLoader.__new__(ChunkedDataLoader)
        chunks = loader._load_excel_chunked(
            str(tmp_path / "t.xlsx"), "Sheet1", header_row=2, chunk_size=50
        )

        assert len(chunks) == 3
        assert list(chunks[0].columns) == ["id", "amount"]
        assert chunks[0].index.tolist() == list(range(0, 50))
        assert chunks[1].index.tolist() == list(range(50, 100))
        assert chunks[2].index.tolist() == list(range(100, 120))
        assert chunks[2].iloc[-1]["id"] == 119

    def test_trailing_empty_rows_dropped(self, tmp_path):
        """数据后存在仅格式化的空行时，尾部空行必须被丢弃（与 read_excel 语义一致）。"""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["id", "name"])
        for i in range(150):
            ws.append([i, f"n{i}"])
        for r in range(153, 156):
            for c in range(1, 3):
                ws.cell(row=r, column=c).font = Font(bold=True)
        wb.save(tmp_path / "t.xlsx")
        wb.close()

        loader = ChunkedDataLoader.__new__(ChunkedDataLoader)
        chunks = loader._load_excel_chunked(
            str(tmp_path / "t.xlsx"), "Sheet1", header_row=0, chunk_size=100
        )

        assert len(chunks) == 2
        assert len(chunks[0]) == 100
        assert len(chunks[1]) == 50
        assert chunks[1].index.tolist() == list(range(100, 150))
