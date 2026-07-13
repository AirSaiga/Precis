"""
@fileoverview CSV/Excel 加载器单元测试

测试 CSVLoader 和 ExcelLoader 的 load/validate/preview 方法。
通过先构造合法相对路径再覆盖 path 属性绕过 Windows 绝对路径校验。
"""

import os
import sys

_project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)


import pandas as pd
import pytest

from app.shared.core.data_source.loaders.base import DataLoadError
from app.shared.core.data_source.loaders.csv_loader import CSVLoader
from app.shared.core.data_source.loaders.excel_loader import ExcelLoader
from app.shared.core.data_source.specs.csv_source import CSVSourceSpec
from app.shared.core.data_source.specs.excel_source import ExcelSourceSpec


def _make_csv_spec(real_path: str, **overrides):
    spec = CSVSourceSpec(path="test.csv", mode="relative", **overrides)
    spec.path = real_path
    return spec


def _make_excel_spec(real_path: str, **overrides):
    spec = ExcelSourceSpec(path="test.xlsx", mode="relative", **overrides)
    spec.path = real_path
    return spec


class TestCSVLoader:
    def test_load_basic(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file))
        loader = CSVLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["a", "b"]
        assert len(df) == 2

    def test_load_no_header(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("1,2\n3,4\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file), header_enabled=False)
        loader = CSVLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["col_0", "col_1"]

    def test_load_skip_rows(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("skip\na,b\n1,2\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file), skip_rows=1)
        loader = CSVLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["a", "b"]
        assert len(df) == 1

    def test_load_nrows(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("a,b\n1,2\n3,4\n5,6\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file), nrows=2)
        loader = CSVLoader(spec)
        df = loader.load()
        assert len(df) == 2

    def test_load_file_not_found(self):
        spec = _make_csv_spec("D:\\nonexistent.csv")
        loader = CSVLoader(spec)
        with pytest.raises(DataLoadError):
            loader.load()

    def test_resolve_encoding_no_detection(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("a\n1\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file), encoding_detection=False)
        loader = CSVLoader(spec)
        assert loader._resolve_encoding() == "utf-8"

    def test_resolve_encoding_with_detection(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("a\n1\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file), encoding_detection=True)
        loader = CSVLoader(spec)
        assert loader._resolve_encoding() == "utf-8"

    def test_validate_exists(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("a,b\n1,2\n", encoding="utf-8")
        spec = _make_csv_spec(str(csv_file))
        loader = CSVLoader(spec)
        errors = loader.validate()
        assert errors == []

    def test_validate_not_exists(self):
        spec = _make_csv_spec("D:\\nonexistent.csv")
        loader = CSVLoader(spec)
        errors = loader.validate()
        assert any("不存在" in e for e in errors)

    def test_validate_wrong_extension(self, tmp_path):
        """B19 回归：扩展名不符只是警告（记录日志），不再作为错误返回。
        过去 warning 被加入 errors 列表，调用方按 errors 非空判定为加载失败。"""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("a,b\n1,2\n", encoding="utf-8")
        spec = _make_csv_spec(str(txt_file))
        loader = CSVLoader(spec)
        errors = loader.validate()
        # 非 .csv 扩展名不应阻止加载，errors 应为空
        assert errors == []


class TestExcelLoader:
    def test_load_basic(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        df_orig = pd.DataFrame({"a": [1, 3], "b": [2, 4]})
        df_orig.to_excel(xlsx_file, index=False)
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["a", "b"]
        assert len(df) == 2

    def test_load_sheet_name(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        with pd.ExcelWriter(xlsx_file) as writer:
            pd.DataFrame({"x": [1]}).to_excel(writer, sheet_name="Sheet1", index=False)
            pd.DataFrame({"y": [2]}).to_excel(writer, sheet_name="Sheet2", index=False)
        spec = _make_excel_spec(str(xlsx_file), sheet="Sheet2")
        loader = ExcelLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["y"]

    def test_load_no_header(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        df_orig = pd.DataFrame({0: [1, 3], 1: [2, 4]})
        df_orig.to_excel(xlsx_file, index=False, header=False)
        spec = _make_excel_spec(str(xlsx_file), header_enabled=False)
        loader = ExcelLoader(spec)
        df = loader.load()
        assert list(df.columns) == ["col_0", "col_1"]

    def test_load_file_not_found(self):
        spec = _make_excel_spec("D:\\nonexistent.xlsx")
        loader = ExcelLoader(spec)
        with pytest.raises(DataLoadError):
            loader.load()

    def test_validate_exists(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        pd.DataFrame({"a": [1]}).to_excel(xlsx_file, index=False)
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        errors = loader.validate()
        assert errors == []

    def test_validate_not_exists(self):
        spec = _make_excel_spec("D:\\nonexistent.xlsx")
        loader = ExcelLoader(spec)
        errors = loader.validate()
        assert any("不存在" in e for e in errors)

    def test_validate_wrong_extension(self, tmp_path):
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("hello", encoding="utf-8")
        spec = _make_excel_spec(str(txt_file))
        loader = ExcelLoader(spec)
        errors = loader.validate()
        assert any("不支持" in e for e in errors)

    def test_preview(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        pd.DataFrame({"a": list(range(20))}).to_excel(xlsx_file, index=False)
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.preview(nrows=5)
        assert len(df) == 5

    def test_merged_cell_single_column_fill(self, tmp_path):
        """单列纵向合并区域：区域内的 NaN 应被合并值填充。

        场景：category 列合并（A2:A4 值 Food），id 列每行有值（保证 pandas 读出 3 行）。
        """
        from openpyxl import Workbook

        xlsx_file = tmp_path / "merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="category")
        ws.cell(row=1, column=2, value="id")
        ws.cell(row=2, column=1, value="Food")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=2, value=2)
        ws.cell(row=4, column=2, value=3)
        ws.merge_cells("A2:A4")
        wb.save(xlsx_file)
        wb.close()
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.load()
        # category 合并值应向下填充到 3 个数据行
        assert df["category"].tolist() == ["Food", "Food", "Food"]
        assert df["id"].tolist() == [1, 2, 3]

    def test_merged_cell_multi_column(self, tmp_path):
        """多列合并区域（A2:B4）：两列都应被填充。"""
        from openpyxl import Workbook

        xlsx_file = tmp_path / "merged2.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="col_a")
        ws.cell(row=1, column=2, value="col_b")
        ws.cell(row=1, column=3, value="id")
        ws.cell(row=2, column=1, value="X")
        ws.cell(row=2, column=2, value="Y")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=3, column=3, value=2)
        ws.cell(row=4, column=3, value=3)
        ws.merge_cells("A2:A4")
        ws.merge_cells("B2:B4")
        wb.save(xlsx_file)
        wb.close()
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.load()
        assert df["col_a"].tolist() == ["X", "X", "X"]
        assert df["col_b"].tolist() == ["Y", "Y", "Y"]

    def test_merged_cell_outside_data_range_skipped(self, tmp_path):
        """合并区域起始行超出 df 行数时，应安全跳过不报错。"""
        from openpyxl import Workbook

        xlsx_file = tmp_path / "edge.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="a")
        ws.cell(row=2, column=1, value=1)
        ws.merge_cells("A2:A10")
        wb.save(xlsx_file)
        wb.close()
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.load()  # 不应抛异常
        assert "a" in df.columns

    def test_merged_cell_fill_preserves_existing_values(self, tmp_path):
        """合并区域内已有非空值的单元格不应被填充覆盖（回溯逻辑正确）。

        场景：A2:A4 合并，但 A3 有独立值（非合并起点）。
        注意：openpyxl 合并区域内只有左上角保留值，其余被清空，
        所以这里测试的是：填充从 A2 向下，A3 若在 read_excel 后仍为 NaN 则填 A2 的值。
        """
        from openpyxl import Workbook

        xlsx_file = tmp_path / "keep.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="cat")
        ws.cell(row=1, column=2, value="id")
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value=1)
        ws.cell(row=3, column=2, value=2)
        ws.cell(row=4, column=2, value=3)
        ws.merge_cells("A2:A4")
        wb.save(xlsx_file)
        wb.close()
        spec = _make_excel_spec(str(xlsx_file))
        loader = ExcelLoader(spec)
        df = loader.load()
        # A2:A4 合并后 read_excel 得到 [A, NaN, NaN]，填充后应全为 A
        # 关键：id 列（非合并列）的已有值不被 touched
        assert df["cat"].tolist() == ["A", "A", "A"]
        assert df["id"].tolist() == [1, 2, 3]
