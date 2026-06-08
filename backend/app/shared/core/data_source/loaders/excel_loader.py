"""
@fileoverview Excel 数据源加载器模块

功能概述:
- 加载 .xlsx / .xls 格式 Excel 文件为 pandas DataFrame
- 支持单 sheet 加载（load）和多 sheet 批量加载（load_multi_sheet）
- 空文件提前检查，给出清晰的 DataLoadError（B13）
- header_row 配置错误检测：若 header_row 之后无数据行则输出警告（B10）
- 使用 openpyxl 前向填充合并单元格，消除 NotNull/Unique 假阳性（B7）
- 支持 dtype_inference、skip_rows、nrows 等参数（B9）

架构设计:
- 继承 DataSourceLoader[ExcelSourceSpec]，通过注册表自动发现
- load() 针对单 sheet，返回单个 DataFrame
- load_multi_sheet() 针对项目多表场景，返回 {schema_id: DataFrame}
- _apply_merged_cell_fill() 为 openpyxl 级别的合并单元格填充逻辑

输入示例:
    spec = ExcelSourceSpec(
        path="data/users.xlsx",
        sheet="Sheet1",
        header_row=0,
        dtype_inference=True
    )
    loader = ExcelLoader(spec)

输出示例:
    df = loader.load()
    # 返回 pandas.DataFrame，合并单元格已按 Excel 显示值填充
    # 若 sheet 不存在则抛出 DataLoadError（B8）
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd

from ..specs.excel_source import ExcelSourceSpec
from .base import DataLoadError, DataSourceLoader
from .registry import register_loader

logger = logging.getLogger(__name__)


@register_loader("excel")
class ExcelLoader(DataSourceLoader[ExcelSourceSpec]):
    """
    @classdesc Excel 文件加载器

    支持 .xlsx 和 .xls 格式的 Excel 文件。
    使用 pandas.read_excel 进行读取。

    支持两种模式：
    - 单 sheet 加载：通过 load() 返回单个 DataFrame
    - 多 sheet 批量加载：通过 load_multi_sheet() 返回 {sheet_name: DataFrame}
    """

    spec_class = ExcelSourceSpec

    def load(self) -> pd.DataFrame:
        """
        @methoddesc 加载 Excel 文件并返回 DataFrame。

        根据 spec 中的 sheet 名称或索引读取指定工作表，
        支持 header_row、skip_rows、nrows 等配置。
        读取后会应用合并单元格前向填充（仅 openpyxl 引擎）。

        Returns:
            加载的 DataFrame

        Raises:
            DataLoadError: 文件不存在、为空、加载失败或 sheet 不存在时抛出

        示例:
            >>> spec = ExcelSourceSpec(path="data.xlsx", sheet="Sheet1", header_row=0)
            >>> loader = ExcelLoader(spec)
            >>> df = loader.load()
        """
        try:
            # 空文件提前检查，给出清晰错误（B13）
            path = Path(self.spec.path)
            if path.exists() and path.stat().st_size == 0:
                raise DataLoadError(f"Excel 文件为空: {self.spec.path}", self.spec)

            read_kwargs = self._build_read_kwargs()

            df = pd.read_excel(self.spec.path, **read_kwargs)

            if not self.spec.header_enabled:
                df.columns = [f"col_{i}" for i in range(len(df.columns))]

            # header_row 配置错误导致空数据时给出警告（B10）
            if len(df) == 0 and self.spec.header_row > 0:
                logger.warning(
                    f"Excel 表 '{self.spec.sheet or self.spec.sheet_index}' "
                    f"在 header_row={self.spec.header_row} 之后没有数据行，"
                    f"请检查 header_row 配置是否正确"
                )

            df = self._apply_merged_cell_fill(df, self.spec.sheet, self.spec.header_row)
            return df

        except FileNotFoundError as e:
            raise DataLoadError(f"文件不存在: {self.spec.path}", self.spec, e)
        except DataLoadError:
            raise
        except Exception as e:
            raise DataLoadError(f"Excel 加载失败: {e}", self.spec, e)

    def load_multi_sheet(
        self,
        sheet_configs: dict[str, dict[str, Any]],
    ) -> dict[str, pd.DataFrame]:
        """
        @methoddesc 批量加载多个 sheet。

        :param sheet_configs: 字典，键为 schema_id，值包含:
            - sheet_name: sheet 名称
            - header_row: 表头行号
        :return: 字典，键为 schema_id，值为对应的 DataFrame
        """
        if not sheet_configs:
            return {}

        sheet_names = list({cfg["sheet_name"] for cfg in sheet_configs.values() if cfg.get("sheet_name")})

        if not sheet_names:
            return {}

        try:
            loaded_sheets: dict = pd.read_excel(
                self.spec.path,
                sheet_name=sheet_names,
                header=None,
                engine=self.spec.engine,
                engine_kwargs={"data_only": True},
            )

            results: dict[str, pd.DataFrame] = {}
            for schema_id, cfg in sheet_configs.items():
                sheet_name = cfg.get("sheet_name")
                header_row = cfg.get("header_row", 0)
                skip_rows = cfg.get("skip_rows", 0)
                nrows = cfg.get("nrows")
                dtype_inference = cfg.get("dtype_inference", True)

                if not sheet_name:
                    continue
                # 缺失 sheet 时显式报错，避免静默跳过导致假阴性（B8）
                if sheet_name not in loaded_sheets:
                    raise DataLoadError(
                        f"Sheet '{sheet_name}' 不存在于文件 {self.spec.path} 中",
                        self.spec,
                    )

                df = loaded_sheets[sheet_name]
                effective_header = header_row + skip_rows
                df.columns = df.iloc[effective_header]
                df = df.drop(index=range(effective_header + 1)).reset_index(drop=True)

                # 应用 nrows 限制（B9）
                if nrows is not None:
                    df = df.head(nrows)

                # 应用 dtype_inference（B9）
                if not dtype_inference:
                    df = df.astype(str)

                # header_row 配置错误导致空数据时给出警告（B10）
                if len(df) == 0:
                    logger.warning(
                        f"Sheet '{sheet_name}' 在 header_row={header_row} 之后没有数据行，"
                        f"请检查 header_row 配置是否正确"
                    )
                df = self._apply_merged_cell_fill(df, sheet_name, effective_header)
                results[schema_id] = df

            return results

        except FileNotFoundError as e:
            raise DataLoadError(f"文件不存在: {self.spec.path}", self.spec, e)
        except Exception as e:
            raise DataLoadError(f"Excel 多 sheet 加载失败: {e}", self.spec, e)

    def _build_read_kwargs(self) -> dict[str, Any]:
        """
        @methoddesc 构造 pandas.read_excel 的参数字典

        业务用途:
        - 根据 self.spec 配置（header 行、engine、dtype、sheet 索引、skip_rows、nrows）组装参数
        - engine_kwargs.data_only=True 用于读取公式结果而非公式本身

        返回:
            可直接传入 pd.read_excel 的参数字典
        """
        read_kwargs: dict[str, Any] = {
            "header": self.spec.header_row if self.spec.header_enabled else None,
            "engine": self.spec.engine,
            "dtype": None if self.spec.dtype_inference else str,
            "engine_kwargs": {"data_only": True},
        }

        if self.spec.sheet:
            read_kwargs["sheet_name"] = self.spec.sheet
        else:
            read_kwargs["sheet_name"] = self.spec.sheet_index

        if self.spec.skip_rows > 0:
            read_kwargs["skiprows"] = self.spec.skip_rows
        if self.spec.nrows:
            read_kwargs["nrows"] = self.spec.nrows

        return read_kwargs

    def _apply_merged_cell_fill(
        self, df: pd.DataFrame, sheet_name: str | None = None, header_row: int = 0
    ) -> pd.DataFrame:
        """@methoddesc 对合并单元格进行前向填充，避免 NotNull/Unique 误报（B7）。

        使用 openpyxl 检测合并单元格区域，然后仅对区域内的 NaN 进行填充。
        """
        if self.spec.engine != "openpyxl":
            return df
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.spec.path, data_only=True)
            try:
                if sheet_name is None:
                    sheet_name = self.spec.sheet or wb.sheetnames[0]
                ws = wb[sheet_name]

                df_filled = df.copy()
                for merged in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = (
                        merged.min_row,
                        merged.min_col,
                        merged.max_row,
                        merged.max_col,
                    )
                    # openpyxl 是 1-based；pandas DataFrame row 0 对应 Excel row (header_row + 2)
                    start_df_row = min_row - header_row - 2
                    end_df_row = max_row - header_row - 2
                    start_df_col = min_col - 1
                    end_df_col = max_col - 1

                    if start_df_row < 0 or start_df_col < 0:
                        continue
                    if start_df_row >= len(df_filled):
                        continue

                    for col_idx in range(start_df_col, min(end_df_col + 1, len(df_filled.columns))):
                        for row_idx in range(start_df_row, min(end_df_row + 1, len(df_filled))):
                            if pd.isna(df_filled.iloc[row_idx, col_idx]):
                                for back_idx in range(row_idx - 1, start_df_row - 1, -1):
                                    if back_idx < 0:
                                        break
                                    if not pd.isna(df_filled.iloc[back_idx, col_idx]):
                                        df_filled.iloc[row_idx, col_idx] = df_filled.iloc[back_idx, col_idx]
                                        break
                return df_filled
            finally:
                wb.close()
        except Exception:
            return df

    def validate(self) -> list[str]:
        """
        @methoddesc 验证 Excel 文件配置和文件本身。

        检查项：
        - 文件是否存在
        - 文件扩展名是否为 .xlsx 或 .xls
        - 文件大小是否超过 100MB（发出警告）

        Returns:
            错误信息列表，空列表表示验证通过

        示例:
            >>> errors = loader.validate()
            >>> if errors:
            ...     print("验证失败:", errors)
        """
        errors = []
        path = Path(self.spec.path)

        if not path.exists():
            errors.append(f"文件不存在: {self.spec.path}")
            return errors

        ext = path.suffix.lower()
        if ext not in [".xlsx", ".xls"]:
            errors.append(f"不支持的 Excel 格式: {ext}")

        size_mb = path.stat().st_size / (1024 * 1024)
        if size_mb > 100:
            errors.append(f"警告: 文件较大 ({size_mb:.1f}MB)")

        return errors

    def preview(self, nrows: int = 10) -> pd.DataFrame:
        """
        @methoddesc 预览 Excel 文件的前 n 行数据。

        通过限制 nrows 参数快速加载文件头部数据，
        比加载完整文件更高效。同样会应用合并单元格填充。

        Args:
            nrows: 要预览的行数，默认为 10

        Returns:
            包含前 n 行数据的 DataFrame

        示例:
            >>> df = loader.preview(nrows=5)
            >>> print(df.head())
        """
        try:
            read_kwargs = {
                "nrows": nrows,
                "header": self.spec.header_row if self.spec.header_enabled else None,
                "engine": self.spec.engine,
                "engine_kwargs": {"data_only": True},
            }

            if self.spec.sheet:
                read_kwargs["sheet_name"] = self.spec.sheet
            else:
                read_kwargs["sheet_name"] = self.spec.sheet_index

            df = pd.read_excel(self.spec.path, **read_kwargs)

            if not self.spec.header_enabled:
                df.columns = [f"col_{i}" for i in range(len(df.columns))]

            df = self._apply_merged_cell_fill(df, self.spec.sheet, self.spec.header_row)
            return df

        except Exception:
            return super().preview(nrows)
